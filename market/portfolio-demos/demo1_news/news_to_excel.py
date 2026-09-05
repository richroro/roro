# -*- coding: utf-8 -*-
"""
데모 1) 뉴스 자동 수집 → 엑셀 정리
------------------------------------------------
특정 뉴스 사이트(RSS)에서 최신 기사 제목/링크/날짜를 자동으로 긁어와
보기 좋게 정리된 엑셀 파일로 저장합니다.

포트폴리오용 데모 스크립트 - 실제 고객 작업도 이 구조를 그대로 응용합니다.
실행:  python news_to_excel.py
결과:  뉴스_수집결과_YYYYMMDD.xlsx
"""
import sys
import datetime
import xml.etree.ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 수집 대상 (공개 RSS 피드 - 누구나 구독 가능하도록 언론사가 제공하는 공식 피드)
RSS_FEEDS = [
    ("연합뉴스-경제", "https://www.yna.co.kr/rss/economy.xml"),
    ("한겨레-경제", "https://www.hani.co.kr/rss/economy/"),
]

SAMPLE_FALLBACK = [
    ("경제", "코스피, 외국인 매수세에 2,700선 회복", "https://example.com/news/1", "2026-08-16"),
    ("경제", "원/달러 환율 소폭 하락… 1,320원대 진입", "https://example.com/news/2", "2026-08-16"),
    ("경제", "반도체 수출 3개월 연속 증가세", "https://example.com/news/3", "2026-08-16"),
    ("경제", "기준금리 동결 전망 우세… 시장 안도", "https://example.com/news/4", "2026-08-16"),
    ("경제", "국제유가 배럴당 78달러 안팎 등락", "https://example.com/news/5", "2026-08-16"),
]


def fetch_rss(name, url, limit=10):
    """RSS 피드에서 기사 제목/링크/날짜를 추출한다."""
    rows = []
    try:
        res = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        res.raise_for_status()
        root = ET.fromstring(res.content)
        for item in root.iter("item"):
            title = (item.findtext("title") or "").strip()
            link = (item.findtext("link") or "").strip()
            pub = (item.findtext("pubDate") or "").strip()
            if title:
                rows.append((name, title, link, pub))
            if len(rows) >= limit:
                break
    except Exception as e:
        print(f"  [알림] {name} 수집 실패({e}) → 건너뜀")
    return rows


def build_excel(rows, path):
    """수집한 데이터를 서식이 적용된 엑셀로 저장한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "뉴스수집결과"

    headers = ["출처", "제목", "링크", "발행일"]
    header_fill = PatternFill("solid", fgColor="2E5D9F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for r in rows:
        ws.append(list(r))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [16, 60, 45, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    # 요약 시트
    ws2 = wb.create_sheet("수집요약")
    ws2["A1"] = "수집 일시"
    ws2["B1"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A2"] = "총 기사 수"
    ws2["B2"] = len(rows)
    ws2["A1"].font = ws2["A2"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30

    wb.save(path)


def main():
    print("=" * 46)
    print("  뉴스 자동 수집 → 엑셀 정리  [데모]")
    print("=" * 46)
    all_rows = []
    for name, url in RSS_FEEDS:
        print(f"- 수집 중: {name}")
        all_rows.extend(fetch_rss(name, url))

    if not all_rows:
        print("  네트워크 미연결/차단 → 샘플 데이터로 대체 생성")
        all_rows = SAMPLE_FALLBACK

    today = datetime.datetime.now().strftime("%Y%m%d")
    out = f"뉴스_수집결과_{today}.xlsx"
    build_excel(all_rows, out)
    print("-" * 46)
    print(f"  완료! 총 {len(all_rows)}건 → '{out}' 저장")
    print("=" * 46)


if __name__ == "__main__":
    main()
