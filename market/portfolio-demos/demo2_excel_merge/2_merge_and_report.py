# -*- coding: utf-8 -*-
"""
데모 2) 엑셀 여러 파일 자동 병합 + 집계 리포트 (After)
------------------------------------------------
input_files/ 폴더의 지점별 매출 파일(여러 개)을 자동으로 찾아
1) 하나의 통합 시트로 합치고
2) 지점별/상품별 매출 집계 리포트까지 자동 생성합니다.

파일이 5개든 50개든 코드 수정 없이 폴더에 넣기만 하면 됩니다.
실행:  python 2_merge_and_report.py
결과:  통합매출_리포트.xlsx
"""
import os
import glob
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

BASE = os.path.dirname(__file__)
IN_DIR = os.path.join(BASE, "input_files")
OUT = os.path.join(BASE, "통합매출_리포트.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="2E5D9F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def main():
    files = sorted(glob.glob(os.path.join(IN_DIR, "매출_*.xlsx")))
    if not files:
        print("input_files 폴더에 매출 파일이 없습니다. 먼저 1_make_sample_files.py 실행하세요.")
        return

    print(f"발견한 파일 {len(files)}개를 병합합니다...")

    merged = []               # (지점, 상품, 수량, 단가, 매출)
    by_branch = defaultdict(int)
    by_product = defaultdict(int)

    for f in files:
        branch = os.path.splitext(os.path.basename(f))[0].replace("매출_", "")
        wb = load_workbook(f)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            name, qty, price = row[0], int(row[1]), int(row[2])
            amount = qty * price
            merged.append((branch, name, qty, price, amount))
            by_branch[branch] += amount
            by_product[name] += amount
        print(f"  + {branch}: {ws.max_row - 1}개 상품")

    out = Workbook()

    # 1) 통합 원본 시트
    ws1 = out.active
    ws1.title = "통합원본"
    ws1.append(["지점", "상품명", "판매수량", "단가", "매출액"])
    style_header(ws1, 5)
    for r in merged:
        ws1.append(list(r))
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=5):
        for c in row:
            c.border = BORDER
    total = sum(r[4] for r in merged)
    ws1.append(["합계", "", "", "", total])
    for c in ws1[ws1.max_row]:
        c.fill = TOTAL_FILL
        c.font = Font(bold=True)
    for col, w in zip("ABCDE", [10, 14, 10, 8, 14]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # 2) 지점별 집계 + 차트
    ws2 = out.create_sheet("지점별집계")
    ws2.append(["지점", "매출액"])
    style_header(ws2, 2)
    for b, amt in sorted(by_branch.items(), key=lambda x: -x[1]):
        ws2.append([b, amt])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    chart = BarChart()
    chart.title = "지점별 매출"
    chart.type = "col"
    data = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(by_branch))
    cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(by_branch))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws2.add_chart(chart, "D2")

    # 3) 상품별 집계
    ws3 = out.create_sheet("상품별집계")
    ws3.append(["상품명", "매출액"])
    style_header(ws3, 2)
    for p, amt in sorted(by_product.items(), key=lambda x: -x[1]):
        ws3.append([p, amt])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 14

    out.save(OUT)
    print("-" * 40)
    print(f"완료! {len(files)}개 파일 → '통합매출_리포트.xlsx'")
    print(f"  · 통합 행: {len(merged)}행")
    print(f"  · 총 매출: {total:,}원")
    print(f"  · 시트: 통합원본 / 지점별집계(+차트) / 상품별집계")


if __name__ == "__main__":
    main()
