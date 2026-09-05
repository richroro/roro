# -*- coding: utf-8 -*-
"""
데모 3) 매일 자동 실행 - 일일 리포트 자동 생성
------------------------------------------------
윈도우 작업 스케줄러에 등록하면 매일 정해진 시간에 자동 실행되어
'그날 날짜'가 찍힌 리포트 엑셀을 자동으로 만들어 냅니다.
사람이 손대지 않아도 매일 결과 파일이 쌓입니다.

수동 실행:  python daily_report.py
자동 실행:  run_daily.bat 을 작업 스케줄러에 등록 (README 참고)
"""
import os
import datetime
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(BASE, "auto_reports")
os.makedirs(OUT_DIR, exist_ok=True)

now = datetime.datetime.now()
stamp = now.strftime("%Y%m%d")


def make_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "일일현황"

    ws["A1"] = f"일일 자동 리포트 ({now.strftime('%Y-%m-%d %H:%M')})"
    ws["A1"].font = Font(bold=True, size=13, color="2E5D9F")
    ws.merge_cells("A1:C1")

    ws.append([])
    ws.append(["항목", "값", "비고"])
    for c in ws[3]:
        c.fill = PatternFill("solid", fgColor="2E5D9F")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    # 데모용 지표 (실제 작업에서는 매출/방문자/재고 등 실데이터가 들어갑니다)
    rows = [
        ("방문자 수", random.randint(800, 1500), "전일 대비 자동 집계"),
        ("주문 건수", random.randint(40, 120), ""),
        ("매출액(원)", f"{random.randint(1500000, 4000000):,}", ""),
        ("문의 접수", random.randint(3, 20), ""),
    ]
    for r in rows:
        ws.append(list(r))

    for col, w in zip("ABC", [16, 16, 22]):
        ws.column_dimensions[col].width = w

    out = os.path.join(OUT_DIR, f"일일리포트_{stamp}.xlsx")
    wb.save(out)
    return out


def append_log(path):
    log = os.path.join(BASE, "실행로그.txt")
    with open(log, "a", encoding="utf-8") as f:
        f.write(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] 자동 생성 완료 -> {os.path.basename(path)}\n")


if __name__ == "__main__":
    out = make_report()
    append_log(out)
    print(f"[{now.strftime('%H:%M:%S')}] 리포트 자동 생성 완료: {os.path.basename(out)}")
    print(f"  저장 위치: {out}")
